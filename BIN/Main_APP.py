#!/usr/bin/env python
# -*- coding:utf-8 -*-

import tkcalendar as tkca
import tkinter.ttk as ttk
import os
from tkinter import *
from tkinter.font import Font
from tkinter.ttk import *
# Usage:showinfo/warning/error,askquestion/okcancel/yesno/retrycancel
from tkinter.messagebox import *
# Usage:f=tkFileDialog.askopenfilename(initialdir='E:/Python')
import tkinter.filedialog as tkFileDialog
import tkinter.simpledialog as tkSimpleDialog  # askstring()
from interface_main import *
# from calendar_module import *
# from tkcalendar import Calendar
import datetime
import openpyxl as xl
import win32com.client
import pythoncom
import threading
import time
from language_dict import *
from interface_CustomerList import Customer_ui
from interface_PersonList import Person_ui


class Connect_Database():
    def __init__(self, LAN):
        self.LAN = LAN
        if self.LAN == 'English':
            self.m = EngDict_Msg
        elif self.LAN == '日本語':
            self.m = JpnDict_Msg
        elif self.LAN == 'ไทย':
            self.m = ThaDict_Msg

        self.db_name = 'Main_DB.xlsx'
        self.wb = xl.load_workbook(self.db_name)
        self.ws = None

    def get_customer_names(self):
        self.ws = self.wb['Customer_Names']
        re = []
        for each_name in self.ws:
            re.append(each_name[0].value + str(each_name[1].value))
        re.sort(key=str.lower)
        return re

    def add_customer_name(self, new_cus):
        self.new_cus = new_cus
        self.ws = self.wb['Customer_Names']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return

        for i in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=1).value == self.new_cus:
                showerror(title=self.m['T25'], message=eval(self.m['M25']))
                return

        for i in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=2).value == 1:
                self.ws.cell(row=i, column=2).value = 0
                break

        self.ws.append([self.new_cus, 1])
        self.wb.save('Main_DB.xlsx')
        showinfo(title=self.m['T26'], message=eval(self.m['M26']))

    def delete_customer_name(self, target_cus):
        self.target_cus = target_cus
        self.ws = self.wb['Customer_Names']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return
        counter = 0
        for each_line in self.ws:
            counter += 1
            if each_line[0].value == self.target_cus:
                if each_line[1].value == 1:
                    self.ws.cell(row=1, column=2).value = 1
                self.ws.delete_rows(counter)
                break
        self.wb.save('Main_DB.xlsx')
        showinfo(title=self.m['T27'], message=eval(self.m['M27']))

    def get_stuff_namesPre(self):
        self.ws = self.wb['Stuff_Names']
        re = []
        for each_name in self.ws:
            re.append(each_name[0].value + str(each_name[1].value))
        re.sort(key=str.lower)
        return re

    def add_stuff_name(self, new_name):
        self.new_name = new_name
        self.ws = self.wb['Stuff_Names']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return

        for i in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=1).value == self.new_name:
                if self.new_name[len(self.new_name) - 2:].isnumeric() == False:
                    a = askquestion(title=self.m['T28'], message=eval(self.m['M28']))
                    if a == 'yes':
                        self.new_name = self.new_name + '01'
                    else:
                        showinfo(title=self.m['T29'], message=self.m['M29'])
                        return

                else:
                    a = askquestion(title=self.m['T30'], message=eval(self.m['M30']))
                    if a == 'yes':
                        self.new_name = self.new_name[:len(self.new_name) - 2] + str(
                            int(self.new_name[len(self.new_name) - 2:]) + 1)
                    else:
                        showinfo(title=self.m['T29'], message=self.m['M29'])
                        return

        for i in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=2).value == 1:
                self.ws.cell(row=i, column=2).value = 0
                break

        self.ws.append([self.new_name, 1])
        self.wb.save('Main_DB.xlsx')
        showinfo(title=self.m['T31'], message=eval(self.m['M31']))

    def delete_stuff_name(self, target_name):
        self.target_name = target_name
        self.ws = self.wb['Stuff_Names']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return
        counter = 0
        for each_line in self.ws:
            counter += 1
            if each_line[0].value == self.target_name:
                if each_line[1].value == 1:
                    self.ws.cell(row=1, column=2).value = 1
                self.ws.delete_rows(counter)
                break
        self.wb.save('Main_DB.xlsx')
        showinfo(title=self.m['T32'], message=eval(self.m['M32']))

    def get_IP_Paths(self):
        self.ws = self.wb['Paths']
        re = []
        for i in range(1, 4):
            re.append(self.ws[f'A{i}'].value)
        return re

    def set_MasterPath(self, master_path):
        self.ws = self.wb['Paths']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return -1
        self.ws['A2'].value = master_path
        self.wb.save('Main_DB.xlsx')

    def set_ServerIP(self, server_ip):
        self.ws = self.wb['Paths']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return -1
        self.ws['A1'].value = server_ip
        self.wb.save('Main_DB.xlsx')

    def set_PM_Path(self, pm_path):
        self.ws = self.wb['Paths']
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return -1
        self.ws['A3'].value = pm_path
        self.wb.save('Main_DB.xlsx')

    def mark_default_option(self, type, name):
        try:
            self.wb.save('Main_DB.xlsx')
        except PermissionError:
            showwarning(title=self.m['T24'], message=self.m['M24'])
            return
        if type == 'customer':
            self.ws = self.wb['Customer_Names']

        elif type == 'person':
            self.ws = self.wb['Stuff_Names']

        else:
            return

        for i in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=2).value == 1:
                self.ws.cell(row=i, column=2).value = 0
                break

        for i in range(1, self.ws.max_row + 1):
            if self.ws.cell(row=i, column=1).value == name:
                self.ws.cell(row=i, column=2).value = 1
                break
        self.wb.save('Main_DB.xlsx')


class Connect_Server():
    def __init__(self, LAN, master_path, pm_path, date, customer, project_name, submission, person, mode):
        self.LAN = LAN
        if self.LAN == 'English':
            self.m = EngDict_Msg
        elif self.LAN == '日本語':
            self.m = JpnDict_Msg
        elif self.LAN == 'ไทย':
            self.m = ThaDict_Msg

        self.master_path = master_path
        self.pm_path = pm_path

        self.date = date
        self.customer = customer
        self.project_name = project_name
        self.submission = submission
        self.person = person
        self.mode = mode
        self.project_id = ''
        self.Message = ''
        self.progressmax = 100
        self.progressval = 0


    def master_permission_check(self, project_names):
        pythoncom.CoInitialize()
        self.Message = self.m['M33']
        self.progressval = 50
        self.objExcelFile = win32com.client.Dispatch("Excel.Application")
        try:
            self.objExcelFile.Visible = False
            self.objExcelFile.DisplayAlerts = False
            self.objExcelFile.ScreenUpdating = False
            self.objWorkBook = self.objExcelFile.Workbooks.open(self.master_path)
        except:
            try:
                try:
                    self.objWorkBook.Close(False)
                except:
                    pass
                self.objExcelFile.Quit()
                self.objImportSheet = None
                self.objWorkBook = None
                self.objExcelFile = None
            except:
                pass
            pythoncom.CoUninitialize()
            self.Message = self.m['M34']
            showerror(title=self.m['T35'], message=self.m['M35'])
            return

        found_sheet = False
        self.Message = self.m['M36']
        self.progressval = 100
        for s in range(0, len(self.objWorkBook.WorkSheets)):
            if self.objWorkBook.WorkSheets[s].Name[:3].lower() == 'quo':
                self.objImportSheet = self.objWorkBook.WorkSheets[s]
                found_sheet = True
                break
        if found_sheet == False:
            self.objWorkBook.Save()
            self.objWorkBook.Close(False)
            self.objExcelFile.Quit()
            self.objImportSheet = None
            self.objWorkBook = None
            self.objExcelFile = None
            pythoncom.CoUninitialize()
            self.Message = self.m['M37']
            showerror(title=self.m['T38'], message=self.m['M38'])
            return

        self.objWorkBook.Save()
        self.objWorkBook.Close(False)
        self.objExcelFile.Quit()
        self.objImportSheet = None
        self.objWorkBook = None
        self.objExcelFile = None
        pythoncom.CoUninitialize()

        if project_names == []:
            self.fill_master_excel()
        else:
            self.batch_processing(project_names=project_names)

    def batch_processing(self, project_names):
        self.Message = self.m['M39']
        self.progressval = 0
        pythoncom.CoInitialize()
        self.objExcelFile = win32com.client.Dispatch("Excel.Application")
        try:
            self.objExcelFile.Visible = False
        except:
            pass
        self.objExcelFile.DisplayAlerts = False
        self.objExcelFile.ScreenUpdating = False
        self.objWorkBook = self.objExcelFile.Workbooks.open(self.master_path)

        found_sheet = False
        for s in range(0, len(self.objWorkBook.WorkSheets)):
            if self.objWorkBook.WorkSheets[s].Name[:3].lower() == 'quo':
                self.objImportSheet = self.objWorkBook.WorkSheets[s]
                found_sheet = True
                break
        if found_sheet == False:
            self.objWorkBook.Save()
            self.objWorkBook.Close(False)
            self.objExcelFile.Quit()
            self.objImportSheet = None
            self.objWorkBook = None
            self.objExcelFile = None
            pythoncom.CoUninitialize()
            self.Message = self.m['M40']

            showerror(title=self.m['T38'], message=self.m['M38'])
            return

        self.active_col = 1
        self.active_row = 1

        self.Message = self.m['M41']
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '見積管理番号' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                break
        for i in range(4, self.objImportSheet.UsedRange.Rows.Count + 1):
            if self.objImportSheet.Cells(i, self.active_col).Value == None:
                self.active_row = i
                break

        self.Message = self.m['M42']
        key_columns = ['年月', '得意先', '商品名', '提出', '担当者']
        for key_column in key_columns:
            for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
                if key_column in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                    if key_column == '年月':
                        self.date_col = j
                    elif key_column == '得意先':
                        self.customer_col = j
                    elif key_column == '商品名':
                        self.projectname_col = j
                    elif key_column == '提出':
                        self.submission_col = j
                    else:
                        self.person_col = j
                    break
        self.P_IDs = []
        self.progressmax = len(project_names)
        self.progressval = 0
        for project_name in project_names:
            self.progressval += 1
            for i in range(self.active_row, self.active_row + len(project_names)):
                last_id = self.objImportSheet.Cells(i - 1, self.active_col).Value
                self.project_id = self.ID_increasing(last_ID=last_id, worksheet=self.objImportSheet)
                self.Message = eval(self.m['M43'])

                self.objImportSheet.Cells(i, self.active_col).Value = self.project_id
                self.objImportSheet.Cells(i, self.date_col).Value = self.date
                self.objImportSheet.Cells(i, self.customer_col).Value = self.customer
                self.objImportSheet.Cells(i, self.projectname_col).Value = project_name
                self.objImportSheet.Cells(i, self.submission_col).Value = self.submission
                self.objImportSheet.Cells(i, self.person_col).Value = self.person
                self.P_IDs.append(self.project_id)
                self.active_row += 1
                break

        time.sleep(0.3)
        self.Message = self.m['M44']
        self.objWorkBook.Save()
        self.objWorkBook.Close(False)
        self.objExcelFile.Quit()
        self.objImportSheet = None
        self.objWorkBook = None
        self.objExcelFile = None
        pythoncom.CoUninitialize()

        self.create_batch_folders(project_names=project_names)

    def ID_increasing(self, last_ID, worksheet):
        if last_ID[:3] != 'AKT':
            return 'AKT' + worksheet.Name[len(worksheet.Name) - 2:] + '-0001'
        last_int = int(last_ID[len(last_ID) - 4:])
        last_head = last_ID[:len(last_ID) - 4]
        current_int = str(last_int + 1)
        for x in range(0, 4 - len(current_int)):
            current_int = '0' + current_int
        current_ID = last_head + current_int
        return current_ID

    def fill_master_excel(self):
        self.Message = self.m['M39']
        self.progressval = 0
        pythoncom.CoInitialize()
        self.objExcelFile = win32com.client.Dispatch("Excel.Application")
        try:
            self.objExcelFile.Visible = False
        except:
            pass
        self.objExcelFile.DisplayAlerts = False
        self.objExcelFile.ScreenUpdating = False
        self.objWorkBook = self.objExcelFile.Workbooks.open(self.master_path)

        found_sheet = False
        for s in range(0, len(self.objWorkBook.WorkSheets)):
            if self.objWorkBook.WorkSheets[s].Name[:3].lower() == 'quo':
                self.objImportSheet = self.objWorkBook.WorkSheets[s]
                found_sheet = True
                break
        if found_sheet == False:
            self.objWorkBook.Save()
            self.objWorkBook.Close(False)
            self.objExcelFile.Quit()
            self.objImportSheet = None
            self.objWorkBook = None
            self.objExcelFile = None
            pythoncom.CoUninitialize()
            self.Message = self.m['M40']

            showerror(title=self.m['T38'], message=self.m['M38'])
            return

        self.active_col = 1
        self.active_row = 1

        self.Message = self.m['M41']
        self.progressval = 10
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '見積管理番号' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                break
        for i in range(4, self.objImportSheet.UsedRange.Rows.Count + 1):
            if self.objImportSheet.Cells(i, self.active_col).Value == None:
                self.active_row = i
                break

        last_id = self.objImportSheet.Cells(self.active_row - 1, self.active_col).Value
        self.project_id = self.ID_increasing(last_ID=last_id, worksheet=self.objImportSheet)
        self.objImportSheet.Cells(self.active_row, self.active_col).Value = self.project_id

        self.Message = self.m['M45']
        self.progressval = 20
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '年月' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                self.objImportSheet.Cells(self.active_row, self.active_col).Value = self.date
                break

        self.Message = self.m['M46']
        self.progressval = 40
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '得意先' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                self.objImportSheet.Cells(self.active_row, self.active_col).Value = self.customer
                break

        self.Message = self.m['M47']
        self.progressval = 60
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '商品名' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                self.objImportSheet.Cells(self.active_row, self.active_col).Value = self.project_name
                break

        self.Message = self.m['M48']
        self.progressval = 80
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '提出' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                self.objImportSheet.Cells(self.active_row, self.active_col).Value = self.submission
                break

        self.Message = self.m['M49']
        self.progressval = 100
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '担当者' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                self.objImportSheet.Cells(self.active_row, self.active_col).Value = self.person
                break

        time.sleep(0.3)
        self.Message = self.m['M44']
        self.objWorkBook.Save()
        self.objWorkBook.Close(False)
        self.objExcelFile.Quit()
        self.objImportSheet = None
        self.objWorkBook = None
        self.objExcelFile = None
        pythoncom.CoUninitialize()

        self.create_pm_folder()

    def redo_master_excel(self):
        pythoncom.CoInitialize()
        self.objExcelFile = win32com.client.Dispatch("Excel.Application")
        self.objExcelFile.Visible = False
        self.objExcelFile.DisplayAlerts = False
        self.objExcelFile.ScreenUpdating = False
        self.objWorkBook = self.objExcelFile.Workbooks.open(self.master_path)

        for s in range(0, len(self.objWorkBook.WorkSheets)):
            if self.objWorkBook.WorkSheets[s].Name[:3].lower() == 'quo':
                self.objImportSheet = self.objWorkBook.WorkSheets[s]
                found_sheet = True
                break

        self.active_col = 1
        self.active_row = 1
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            if '見積管理番号' in self.objImportSheet.Cells(4, j).Value.replace(' ', ''):
                self.active_col = j
                break
        for i in range(4, self.objImportSheet.UsedRange.Rows.Count + 1):
            if self.objImportSheet.Cells(i, self.active_col).Value == None:
                self.active_row = i - 1
                break
        for j in range(1, self.objImportSheet.UsedRange.Columns.Count + 1):
            self.objImportSheet.Cells(self.active_row, j).Value = None

        self.objWorkBook.Save()
        self.objWorkBook.Close(False)
        self.objExcelFile.Quit()
        self.objImportSheet = None
        self.objWorkBook = None
        self.objExcelFile = None
        pythoncom.CoUninitialize()

    def analyze_rootname(self, rootname):
        if rootname[len(rootname) - 8:len(rootname) - 2] == 'backup' and rootname[len(rootname) - 2:].isdigit():
            if int(rootname[len(rootname) - 2:]) < 9:
                return rootname[:len(rootname) - 2] + '0' + str(int(rootname[len(rootname) - 2:]) + 1)
            else:
                return rootname[:len(rootname) - 2] + str(int(rootname[len(rootname) - 2:]) + 1)
        else:
            return rootname + ' backup01'

    def create_batch_folders(self, project_names):
        self.progressmax = len(project_names)
        self.progressval = 0
        counter = 0
        for each_name in project_names:
            self.Message = eval(self.m['M50'])

            root_name = self.P_IDs[counter] + '_' + each_name
            counter += 1
            self.progressval = counter
            if os.path.exists(self.pm_path + '/' + root_name):
                temp_name = root_name
                while os.path.exists(self.pm_path + '/' + self.analyze_rootname(rootname=temp_name)):
                    temp_name = self.analyze_rootname(rootname=temp_name)

                showwarning(title=self.m['T51'], message=eval(self.m['M51']))
                root_name = self.analyze_rootname(rootname=temp_name)
            else:
                pass
            root_path = self.pm_path + '/' + root_name
            os.mkdir(path=root_path)

            sub_folders = ['E1_Quotation', 'E2_Source-Data', 'E3_Outsource', 'E4_Feedback', 'E5_To_DTP',
                           'E6_Translation',
                           'E7_Others', 'E8_Finish']
            for sub_folder in sub_folders:
                os.mkdir(path=root_path + '/' + sub_folder)

            root_path = root_path + '/E1_Quotation'

            if self.mode==0:
                sub_folders = ['01_Kou-su from editor', '02_Rough quotation', '03_Latest quotation', 'Analyze Material']
            else:
                sub_folders = ['01_PO', '02_Quo', '03_Quo_sign']

            for sub_folder in sub_folders:
                os.mkdir(path=root_path + '/' + sub_folder)

            if self.mode==0:
                root_path = root_path + '/Analyze Material'
                sub_folders = ['Analyze Report', 'Cost Calculation', 'OS quotation', 'Printing Standard']
                for sub_folder in sub_folders:
                    os.mkdir(path=root_path + '/' + sub_folder)
            else:
                root_path = self.pm_path + '/' + root_name
                root_path = root_path + '/E2_Source-Data'
                sub_folders = ['From AKJ']
                for sub_folder in sub_folders:
                    os.mkdir(path=root_path + '/' + sub_folder)

            self.Message = eval(self.m['M52'])

        self.Message = self.m['M53']
        time.sleep(0.3)

    def create_pm_folder(self):
        self.Message = self.m['M54']
        self.progressval = 0
        root_name = self.project_id + '_' + self.project_name
        if os.path.exists(self.pm_path + '/' + root_name):
            temp_name = root_name
            while os.path.exists(self.pm_path + '/' + self.analyze_rootname(rootname=temp_name)):
                temp_name = self.analyze_rootname(rootname=temp_name)
            ans = askyesno(title=self.m['T55'], message=eval(self.m['M55']))
            if not ans:
                self.Message = self.m['M56']
                self.redo_master_excel()
                self.Message = self.m['M57']
                showinfo(title=self.m['T57'], message=self.m['M57'])
                return
            else:
                root_name = self.analyze_rootname(rootname=temp_name)
        else:
            pass
        root_path = self.pm_path + '/' + root_name
        os.mkdir(path=root_path)

        self.Message = self.m['M58']
        self.progressval = 60
        sub_folders = ['E1_Quotation', 'E2_Source-Data', 'E3_Outsource', 'E4_Feedback', 'E5_To_DTP', 'E6_Translation',
                       'E7_Others', 'E8_Finish']
        for sub_folder in sub_folders:
            os.mkdir(path=root_path + '/' + sub_folder)

        root_path = root_path + '/E1_Quotation'
        if self.mode==0:
            sub_folders = ['01_Kou-su from editor', '02_Rough quotation', '03_Latest quotation', 'Analyze Material']
        else:
            sub_folders = ['01_PO', '02_Quo', '03_Quo_sign']

        for sub_folder in sub_folders:
            os.mkdir(path=root_path + '/' + sub_folder)

        if self.mode==0:
            root_path = root_path + '/Analyze Material'
            sub_folders = ['Analyze Report', 'Cost Calculation', 'OS quotation', 'Printing Standard']
            for sub_folder in sub_folders:
                os.mkdir(path=root_path + '/' + sub_folder)
        else:
            root_path = self.pm_path + '/' + root_name
            root_path = root_path + '/E2_Source-Data'
            sub_folders = ['From AKJ']
            for sub_folder in sub_folders:
                os.mkdir(path=root_path+ '/' + sub_folder)

        self.progressval = 100
        self.Message = eval(self.m['M59'])
        time.sleep(0.3)


class Application(Application_ui, Connect_Database):
    # The class will implement callback function for events and your logical code.
    def __init__(self, LAN, master=None):
        self.LAN = LAN
        if self.LAN == 'English':
            self.m = EngDict_Msg
        elif self.LAN == '日本語':
            self.m = JpnDict_Msg
        elif self.LAN == 'ไทย':
            self.m = ThaDict_Msg
        Application_ui.__init__(self, self.LAN, master)
        try:
            Connect_Database.__init__(self, self.LAN)
        except FileNotFoundError:
            showerror(title=self.m['T1'], message=self.m['M1'])
            top.destroy()
            return
        self.DateTVar.set(datetime.date.today().strftime('%m/%d/%Y'))
        self.LanguageC.bind("<<ComboboxSelected>>", self.language_switch)
        self.ServerT.bind("<Return>", self.IP_confirmed)
        self.refresh_CustomerCombo()
        self.refresh_PersonCombo()
        self.refresh_Paths()

        self.Message = ''
        self.progresscount = 0
        self.progressmax = 100

    def refresh_Paths(self):
        self.paths = self.cd.get_IP_Paths()
        self.ServerTVar.set(self.paths[0])
        self.MasterTVar.set(self.paths[1])
        self.FolderTVar.set(self.paths[2])

    def refresh_CustomerCombo(self):
        self.cd = Connect_Database(self.LAN)
        self.CustomerC['values'] = ()
        self.t1 = self.cd.get_customer_names()
        self.t2 = []
        for each_Cus in self.t1:
            self.t2.append(each_Cus[:len(each_Cus) - 1])
        self.CustomerC['values'] = self.t2
        self.counter2 = -1
        for each_Cus2 in self.t1:
            self.counter2 += 1
            if each_Cus2[len(each_Cus2) - 1:] == '1':
                self.CustomerC.current(self.counter2)
                break

    def refresh_PersonCombo(self):
        self.cd = Connect_Database(self.LAN)
        self.PersonC['values'] = ()
        self.temp1 = self.cd.get_stuff_namesPre()
        self.temp2 = []
        for each_pre in self.temp1:
            self.temp2.append(each_pre[:len(each_pre) - 1])
        self.PersonC['values'] = self.temp2
        self.counter = -1
        for each_pre2 in self.temp1:
            self.counter += 1
            if each_pre2[len(each_pre2) - 1:] == '1':
                self.PersonC.current(self.counter)
                break

    def btnExit_Cmd(self, event=None):
        # Close system
        quit_confirming()

    def lock_buttons(self):
        self.btnDate['state'] = 'disabled'
        self.btnCustomer['state'] = 'disabled'
        self.btnPerson['state'] = 'disabled'
        self.btnServer['state'] = 'disabled'
        self.btnMaster['state'] = 'disabled'
        self.btnFolder['state'] = 'disabled'
        self.btnTest['state'] = 'disabled'
        self.btnGenerate['state'] = 'disabled'
        self.btnBatch['state'] = 'disabled'

    def unlock_buttons(self):
        self.btnDate['state'] = 'enabled'
        self.btnCustomer['state'] = 'enabled'
        self.btnPerson['state'] = 'enabled'
        self.btnServer['state'] = 'enabled'
        self.btnMaster['state'] = 'enabled'
        self.btnFolder['state'] = 'enabled'
        self.btnTest['state'] = 'enabled'
        self.btnGenerate['state'] = 'enabled'
        self.btnBatch['state'] = 'enabled'

    def Main_Generator(self):
        # bind to btnGenerate_Cmd
        #print(self.CheckVar1.get())

        self.lock_buttons()
        check = self.connection_test()
        self.lock_buttons()
        if check == -1:
            self.unlock_buttons()
            return
        self.Message = self.m['M2']
        cs = Connect_Server(LAN=self.LAN, master_path=self.MasterTVar.get(), pm_path=self.FolderTVar.get(),
                            date=self.DateTVar.get(), customer=self.CustomerCVar.get(),
                            project_name=self.ProjectTVar.get().strip(),
                            submission=self.SubmissionCVar.get(), person=self.PersonCVar.get(), mode=self.CheckVar1.get())

        def signal_back():
            while self.th1.is_alive():
                time.sleep(0.03)
                self.Message = cs.Message
                self.progressmax = cs.progressmax
                self.progresscount = cs.progressval

            time.sleep(0.3)

        self.th4 = threading.Thread(target=signal_back)
        self.th4.setDaemon(True)
        self.th4.start()
        cs.master_permission_check(project_names=[])
        self.projectID = cs.project_id
        self.unlock_buttons()

    def btnGenerate_Cmd(self, event=None):
        if not self.ProjectTVar.get().replace(' ', ''):
            showerror(title=self.m['T3'], message=self.m['M3'])
            return
        self.th1 = threading.Thread(target=self.Main_Generator)
        self.th2 = threading.Thread(target=self.gui_update, args=('generate',))
        self.th1.setDaemon(True)
        self.th2.setDaemon(True)
        self.th1.start()
        self.th2.start()

    def gui_update(self, type):
        self.ProgressBar1['maximum'] = self.progressmax
        self.ProgressBar1Var.set(str(self.progresscount))
        while self.th1.is_alive():
            compare = self.Message
            time.sleep(0.03)
            if compare != self.Message:
                self.MessageLVar.set(self.Message)
                self.ProgressBar1['maximum'] = self.progressmax
                self.ProgressBar1Var.set(str(self.progresscount))

        if type == 'test':
            if self.Message == self.m['M4']:
                showinfo(title=self.m['T4'], message=self.m['M4'])
        elif type == 'generate':
            if self.Message == eval(self.m['M5']):
                showinfo(title=self.m['T5'], message=eval(self.m['M5']))
        else:
            if self.Message == self.m['M6']:
                showinfo(title=self.m['T6'], message=self.m['M6'])

    def btnPerson_Cmd(self, event=None):
        class Person(Person_ui):
            # The class will implement callback function for events and your logical code.
            def __init__(self, LAN, master=None):
                self.LAN = LAN
                if self.LAN == 'English':
                    self.m = EngDict_Msg
                elif self.LAN == '日本語':
                    self.m = JpnDict_Msg
                elif self.LAN == 'ไทย':
                    self.m = ThaDict_Msg
                Person_ui.__init__(self, self.LAN, master)
                self.cd = Connect_Database(self.LAN)
                self.refresh_CustomerList()

            def refresh_CustomerList(self):
                self.Clist.delete(0, END)
                re = self.cd.get_stuff_namesPre()
                for each_c in re:
                    self.Clist.insert(END, each_c[:len(each_c) - 1])

            def btnAddC_Cmd(self, event=None):
                new_name = self.InputBoxVar.get()
                if new_name == '':
                    showwarning(title=self.m['T7'], message=self.m['M7'])
                    return
                self.cd.add_stuff_name(new_name=new_name)
                self.refresh_CustomerList()

            def btnDeleteC_Cmd(self, event=None):
                selected_index = self.Clist.curselection()
                if selected_index == '':
                    showwarning(title=self.m['T8'], message=self.m['M8'])
                    return
                target_name = self.Clist.get(selected_index)
                a = askquestion(title=self.m['T9'], message=eval(self.m['M9']))
                if a == 'no':
                    return
                else:
                    self.cd.delete_stuff_name(target_name=target_name)
                    self.refresh_CustomerList()

            def btnResume_Cmd(self, event=None):
                main_reshow()

            def btnOK_Cmd(self, event=None):
                # Selected item: name=self.Clist.get(self.Clist.curselection()[0])
                # print(self.Clist.curselection())
                try:
                    self.cd.mark_default_option(type='person', name=self.Clist.get(self.Clist.curselection()[0]))
                except IndexError:
                    pass
                main_reshow()

        def main_reshow():
            top.deiconify()
            self.refresh_PersonCombo()
            self.top_P.destroy()

        self.top.withdraw()
        self.top_P = Toplevel()
        self.top_P.protocol('WM_DELETE_WINDOW', main_reshow)
        Person(self.LAN, self.top_P).mainloop()

    def btnFolder_Cmd(self, event=None):
        a = tkFileDialog.askdirectory(title=self.m['T10'], initialdir='//' + self.ServerTVar.get() + '/PM_secretB/')
        if a == "":
            return
        check = self.cd.set_PM_Path(pm_path=a)
        if check == -1:
            return
        self.FolderTVar.set(a)

    def btnMaster_Cmd(self, event=None):
        a = tkFileDialog.askopenfilename(title=self.m['T11'],
                                         initialdir='//' + self.ServerTVar.get() + '/Sales_secretB')
        if a == "":
            return
        check = self.cd.set_MasterPath(master_path=a)
        if check == -1:
            return
        self.MasterTVar.set(a)

    def IP_confirmed(self, event):
        def input_check(input_ip):
            for i in range(0, len(input_ip)):
                if input_ip[i:i + 1].isnumeric() == True or input_ip[i:i + 1] == '.':
                    continue
                else:
                    showerror(title=self.m['T12'], message=self.m['M12'])
                    return -1
            return 0

        res = input_check(self.ServerTVar.get())
        if res == -1:
            return
        check = self.cd.set_ServerIP(server_ip=self.ServerTVar.get())
        if check == -1:
            return
        self.ServerT['state'] = 'readonly'

    def btnServer_Cmd(self, event=None):
        self.ServerT['state'] = 'enabled'

    def btnCustomer_Cmd(self, event=None):

        class Customer(Customer_ui):
            # The class will implement callback function for events and your logical code.
            def __init__(self, LAN, master=None):
                self.LAN = LAN
                if self.LAN == 'English':
                    self.m = EngDict_Msg
                elif self.LAN == '日本語':
                    self.m = JpnDict_Msg
                elif self.LAN == 'ไทย':
                    self.m = ThaDict_Msg
                Customer_ui.__init__(self, self.LAN, master)
                self.cd = Connect_Database(self.LAN)
                self.refresh_CustomerList()

            def refresh_CustomerList(self):
                self.Clist.delete(0, END)
                re = self.cd.get_customer_names()
                for each_c in re:
                    self.Clist.insert(END, each_c[:len(each_c) - 1])

            def btnAddC_Cmd(self, event=None):
                new_cus = self.InputBoxVar.get()
                if new_cus == '':
                    showwarning(title=self.m['T7'], message=self.m['M7'])
                    return
                self.cd.add_customer_name(new_cus=new_cus)
                self.refresh_CustomerList()

            def btnDeleteC_Cmd(self, event=None):
                selected_index = self.Clist.curselection()
                if selected_index == '':
                    showwarning(title=self.m['T8'], message=self.m['M8'])
                    return
                target_cus = self.Clist.get(selected_index)
                a = askquestion(title=self.m['T13'], message=eval(self.m['M13']))
                if a == 'no':
                    return
                else:
                    self.cd.delete_customer_name(target_cus=target_cus)
                    self.refresh_CustomerList()

            def btnResume_Cmd(self, event=None):
                main_reshow()

            def btnOK_Cmd(self, event=None):
                # Selected item: name=self.Clist.get(self.Clist.curselection()[0])
                # print(self.Clist.curselection())
                try:
                    self.cd.mark_default_option(type='customer', name=self.Clist.get(self.Clist.curselection()[0]))
                except IndexError:
                    pass
                main_reshow()

        def main_reshow():
            self.top.deiconify()
            self.refresh_CustomerCombo()
            top_C.destroy()

        self.top.withdraw()
        top_C = Toplevel()
        top_C.protocol('WM_DELETE_WINDOW', main_reshow)
        Customer(self.LAN, top_C).mainloop()

    def btnTest_Cmd(self, event=None):
        self.th1 = threading.Thread(target=self.connection_test)
        self.th2 = threading.Thread(target=self.gui_update, args=('test',))
        self.th1.setDaemon(True)
        self.th2.setDaemon(True)
        self.th1.start()
        self.th2.start()

    def print_sel(self, event):
        self.returnVal = self.cal.selection_get()
        self.DateTVar.set(self.returnVal.strftime('%m/%d/%Y'))

    def btnDate_Cmd(self, event=None):
        # Active calendar selector
        self.top1 = Toplevel(self.top)
        self.top1.geometry(f'450x300+400+101')
        try:
            self.cal = tkca.Calendar(self.top1,
                                     font="Arial 14", selectmode='day',
                                     cursor="hand1")  # , year=2018, month=2, day=5)
        except:
            showerror(title='Error', message='No module called Calendar!')
            return
        self.cal.bind("<<CalendarSelected>>", self.print_sel)
        self.cal.pack(fill="both", expand=True)
        ttk.Button(self.top1, text="ok", command=(self.top1.destroy)).pack()

    def language_switch(self, event):
        global LANGUAGE
        LANGUAGE = self.LanguageCVar.get()
        Application.__init__(self, self.LanguageCVar.get())

    def connection_test(self):
        self.lock_buttons()
        self.progresscount = 0
        time.sleep(0.1)
        self.progressmax = 3

        time.sleep(0.1)
        self.Message = self.m['M14']
        self.progresscount = 1
        try:
            if os.path.exists("//" + self.ServerTVar.get() + "/Exchange") == False:
                # if os.path.exists(path='D:\\')==False:
                raise (PermissionError)
        except PermissionError:
            self.Message = self.m['M15']
            showerror(title=self.m['T15'], message=self.m['M15'])
            self.progressmax = 100
            self.progresscount = 0
            time.sleep(0.1)
            self.unlock_buttons()
            return -1

        time.sleep(0.1)
        self.Message = self.m['M16']
        self.progresscount = 2
        try:
            if os.access(self.MasterTVar.get(), os.W_OK) == False:
                raise (PermissionError)
            # os.chdir(self.FolderTVar.get())
        except PermissionError:
            self.Message = self.m['M17']
            showerror(title=self.m['T17'], message=self.m['M17'])
            self.progressmax = 100
            self.progresscount = 0
            time.sleep(0.1)
            self.unlock_buttons()
            return -1

        time.sleep(0.1)
        self.Message = self.m['M18']
        self.progresscount = 3
        try:
            if os.path.exists(self.FolderTVar.get()) == False:
                raise (PermissionError)
        except PermissionError:
            self.Message = self.m['M19']
            showerror(title=self.m['T19'], message=self.m['M19'])
            self.progressmax = 100
            self.progresscount = 0
            time.sleep(0.1)
            self.unlock_buttons()
            return -1
        else:
            self.Message = self.m['M20']
            self.unlock_buttons()
            time.sleep(0.1)

    def Main_Batch(self, project_names):
        # bind to btnBatch_Cmd

        self.lock_buttons()
        check = self.connection_test()
        if check == -1:
            self.unlock_buttons()
            return
        self.lock_buttons()
        self.Message = self.m['M21']
        cs = Connect_Server(LAN=self.LAN, master_path=self.MasterTVar.get(), pm_path=self.FolderTVar.get(),
                            date=self.DateTVar.get(), customer=self.CustomerCVar.get(),
                            project_name=self.ProjectTVar.get().strip(),
                            submission=self.SubmissionCVar.get(), person=self.PersonCVar.get(), mode=self.CheckVar1.get())

        def signal_back():
            while self.th1.is_alive():
                time.sleep(0.03)
                self.Message = cs.Message
                self.progressmax = cs.progressmax
                self.progresscount = cs.progressval

            time.sleep(0.3)

        self.th4 = threading.Thread(target=signal_back)
        self.th4.setDaemon(True)
        self.th4.start()
        cs.master_permission_check(project_names=project_names)
        self.projectID = cs.project_id
        self.unlock_buttons()

    def btnBatch_Cmd(self, event=None):
        def get_project_names(importsheet_path):
            wb = xl.load_workbook(importsheet_path)
            ws = wb.active
            a = []
            for each_line in ws:
                if each_line[0].value == None:
                    continue
                else:
                    a.append(str(each_line[0].value).strip())
            return a

        importsheet_path = tkFileDialog.askopenfilename(title=self.m['T22'],
                                                        filetypes=(('Excel Files', '*xlsx'), ('All files', '*.*')))
        if importsheet_path == "":
            return
        project_names = get_project_names(importsheet_path=importsheet_path)

        self.th1 = threading.Thread(target=self.Main_Batch, args=(project_names,))
        self.th2 = threading.Thread(target=self.gui_update, args=('batch',))
        self.th1.setDaemon(True)
        self.th2.setDaemon(True)
        self.th1.start()
        self.th2.start()


def quit_confirming():
    global LANGUAGE
    if LANGUAGE == 'English':
        m = EngDict_Msg
    elif LANGUAGE == '日本語':
        m = JpnDict_Msg
    elif LANGUAGE == 'ไทย':
        m = ThaDict_Msg
    else:
        m = EngDict_Msg
    answer = askokcancel(title=m['T23'], message=m['M23'])
    if answer == True:
        global wb, ws
        db_name = 'Main_DB.xlsx'
        try:
            wb = xl.load_workbook(db_name)
        except:
            pass
        else:
            ws = wb['Default_Lan']
        if LANGUAGE == 'English':
            ws['A1'].value = 0
        elif LANGUAGE == '日本語':
            ws['A1'].value = 1
        elif LANGUAGE == 'ไทย':
            ws['A1'].value = 2
        else:
            pass
        try:
            wb.save('Main_DB.xlsx')
        except:
            pass
        top.destroy()


if __name__ == "__main__":
    db_name = 'Main_DB.xlsx'
    try:
        wb = xl.load_workbook(db_name)
    except:
        showerror(title='Database Error', message='Unable to start the programme, database is not found!')
        os._exit(0)
    else:
        ws = wb['Default_Lan']
    if ws['A1'].value == 0:
        LANGUAGE = 'English'
    elif ws['A1'].value == 1:
        LANGUAGE = '日本語'
    else:
        LANGUAGE = 'ไทย'

    top = Tk()
    top.protocol('WM_DELETE_WINDOW', quit_confirming)
    APP = Application(LANGUAGE, top)
    APP.mainloop()
